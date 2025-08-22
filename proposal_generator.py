import xml.etree.ElementTree as ET
from datetime import datetime, date, timedelta
import holidays

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, timedelta
import json
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, NextPageTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak, Flowable
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor, black, gray, lightgrey, white
import os
import sys
import csv
import io
import re
import math
import pandas as pd
from models import ProposalItem
from gantt_chart import GanttChartFlowable
# --- MODIFICATION: Added import for openpyxl ---
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:
    messagebox.showerror("Missing Dependency", "The 'openpyxl' library is required to work with Excel files. Please install it using: pip install openpyxl")
    sys.exit()
try:
    from tkcalendar import DateEntry
    HAS_TKCAL = True
except ImportError:
    HAS_TKCAL = False

# --- HELPER FUNCTION FOR PYINSTALLER ---
# This function helps find bundled files (like fonts and logos) when running as an .exe
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

class ProposalGenerator:
    """
    The main application class for the PDF Proposal Generator.
    This version includes a refactored create_pdf method for better
    readability and maintenance.
    """
    def __init__(self, root):
        self.root = root
        self.root.title("Castillo Engineering: Proposal Generator")
        self.root.geometry("1600x900")
        self.root.state('zoomed')
        self.us_holidays = holidays.UnitedStates()

        # --- Initialize data ---
        self.version = tk.StringVar(value="V1") # Version of the proposal
        self.project_name = tk.StringVar(value="Sample Project")
        self.company_name = tk.StringVar(value="Sample Company LLC")
        self.project_start_date = tk.StringVar(value="08/21/25")
        
        self.default_logo_path = resource_path("logo.png")
        self.logo_path = tk.StringVar(value=self.default_logo_path)
        self.client_logo_path = tk.StringVar(value="")
        self.include_gantt = tk.BooleanVar(value=False)
        self.task_counter = 0  # New sequential ID counter
        
        self.template_items = self.create_template_structure()
        self.current_editor = None
        self.drag_data = {"item": None, "index": 0}
        self.item_id_map = {}
        self.link_drag_data = {"start_item_id": None, "last_hover_id": None}
        self.column_drag_data = {}
        
        # --- MODIFICATION: Store last valid start date for reverting changes ---
        self.last_project_start_date = self.project_start_date.get()

        # --- NEW: Make GanttChartFlowable class available to instance ---
        self.GanttChartFlowable = GanttChartFlowable

        self.setup_ui()
        self.populate_tree()
        self.expand_all_items()
    def get_project_end_date(self):
        latest = None
        for item in self.item_id_map.values():
            if item.enabled.get() and item.end_date:
                try:
                    dt = datetime.strptime(item.end_date, "%m/%d/%y")
                    if latest is None or dt > latest:
                        latest = dt
                except ValueError:
                    pass
        return latest.strftime("%m/%d/%y") if latest else None
    def export_to_projectlibre_xml(self):
        MSP_NS = "http://schemas.microsoft.com/project"
        def E(tag):  # element with MSP namespace
            return f"{{{MSP_NS}}}{tag}"

        def iso_dt(x, default_time="08:00:00"):
            """Coerce incoming date/str/datetime -> 'YYYY-MM-DDTHH:MM:SS'."""
            if not x:
                # fall back to project start if available
                try:
                    x = self.project_start_date.get()
                except Exception:
                    x = date.today()
            if isinstance(x, datetime):
                return x.strftime("%Y-%m-%dT%H:%M:%S")
            if isinstance(x, date):
                return datetime(x.year, x.month, x.day, 8, 0, 0).strftime("%Y-%m-%dT%H:%M:%S")
            s = str(x).strip()
            # Try ISO first
            for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
                try:
                    dt = datetime.strptime(s, fmt)
                    if fmt == "%Y-%m-%d":
                        dt = dt.replace(hour=8, minute=0, second=0)
                    return dt.strftime("%Y-%m-%dT%H:%M:%S")
                except Exception:
                    pass
            # Try US format
            for fmt in ("%m/%d/%Y", "%m/%d/%y"):
                try:
                    dt = datetime.strptime(s, fmt).replace(hour=8, minute=0, second=0)
                    return dt.strftime("%Y-%m-%dT%H:%M:%S")
                except Exception:
                    pass
            # Last resort: return as-is if it already looks like yyyy-mm-ddThh:mm:ss
            if "T" in s and len(s) >= 19 and s[4] == "-" and s[7] == "-":
                return s[:19]
            # default now
            return datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

        def to_int(x, d=0):
            try:
                return int(x)
            except Exception:
                return d

        HOURS_PER_DAY = 8  # tune if you use different calendar
        def dur_to_mspdi(duration_days):
            """Duration in MSPDI format as hours, e.g. PT8H0M0S."""
            d = to_int(duration_days, 0)
            hours = max(d, 0) * HOURS_PER_DAY
            return f"PT{hours}H0M0S"

        def classify(item):
            if getattr(item, "children", None):
                return "Summary"
            if getattr(item, "is_milestone", False) or to_int(getattr(item, "duration", 0), 0) == 0:
                return "Milestone"
            return "Task"

        # Try to ensure dates are up-to-date
        try:
            if hasattr(self, "calculate_all_dates") and callable(self.calculate_all_dates):
                self.calculate_all_dates()
        except Exception:
            pass

        # Ask where to save
        filename = filedialog.asksaveasfilename(
            title="Export for ProjectLibre/MS Project (XML)",
            defaultextension=".xml",
            filetypes=[("MS Project XML", "*.xml")]
        )
        if not filename:
            return

        # --- Build MSPDI XML -----------------------------------------------------
        ET.register_namespace("", MSP_NS)
        proj = ET.Element(E("Project"))

        # Basic project meta
        ET.SubElement(proj, E("Name")).text = getattr(self, "project_name", None) and self.project_name.get() or "Exported Project"
        now = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
        ET.SubElement(proj, E("CreationDate")).text = now
        ET.SubElement(proj, E("LastSaved")).text = now
        ET.SubElement(proj, E("ScheduleFromStart")).text = "1"
        ET.SubElement(proj, E("CalendarUID")).text = "1"

        # Default times/units (common MSPDI defaults)
        ET.SubElement(proj, E("DefaultStartTime")).text = "08:00:00"
        ET.SubElement(proj, E("DefaultFinishTime")).text = "17:00:00"
        ET.SubElement(proj, E("MinutesPerDay")).text = str(HOURS_PER_DAY * 60)
        ET.SubElement(proj, E("MinutesPerWeek")).text = str(HOURS_PER_DAY * 5 * 60)
        ET.SubElement(proj, E("DaysPerMonth")).text = "20"

        # Calendars (Standard)
        cals = ET.SubElement(proj, E("Calendars"))
        cal = ET.SubElement(cals, E("Calendar"))
        ET.SubElement(cal, E("UID")).text = "1"
        ET.SubElement(cal, E("Name")).text = "Standard"
        ET.SubElement(cal, E("IsBaseCalendar")).text = "1"
        ET.SubElement(cal, E("BaseCalendarUID")).text = "0"

        # Task container
        tasks_el = ET.SubElement(proj, E("Tasks"))

        # We’ll assign UIDs in traversal order and compute outline numbers
        uid_counter = 1
        id_map = {}        # your item.id -> UID
        outline_stack = [] # for OutlineNumber like 1, 1.1, 1.2 ...

        # Flattened list for a second pass (to add predecessor links)
        flat_items = []

        def walk(items, outline_level=1, prefix_numbers=None):
            nonlocal uid_counter
            if prefix_numbers is None:
                prefix_numbers = []

            for idx, item in enumerate(items or [], start=1):
                # Skip disabled items
                enabled = True
                try:
                    enabled = bool(item.enabled.get())
                except Exception:
                    pass
                if not enabled:
                    # still recurse: enabled grandchildren might exist
                    walk(getattr(item, "children", []), outline_level+1, prefix_numbers + [idx])
                    continue

                row_type = classify(item)

                # Outline number like "1.2.3"
                outline_number = ".".join(map(str, prefix_numbers + [idx]))
                uid = uid_counter
                uid_counter += 1
                id_map[getattr(item, "id", uid)] = uid  # tolerate missing id

                # Dates & Duration
                raw_start = getattr(item, "start_date", None) or getattr(self, "project_start_date", None) and self.project_start_date.get()
                raw_finish = getattr(item, "end_date", None) or raw_start
                start_iso = iso_dt(raw_start)
                finish_iso = iso_dt(raw_finish)
                dur_days = to_int(getattr(item, "duration", 0), 0)
                # Milestone duration must be 0 hours
                if row_type == "Milestone":
                    dur_str = "PT0H0M0S"
                    # For safety, force Start==Finish
                    finish_iso = start_iso
                elif row_type == "Summary":
                    # Summary duration can be left 0; ProjectLibre rolls up
                    dur_str = dur_to_mspdi(0)
                    # Optional: blank start/finish is allowed; but keep ISO so it displays
                else:
                    dur_str = dur_to_mspdi(dur_days)

                # Build <Task>
                t = ET.SubElement(tasks_el, E("Task"))
                ET.SubElement(t, E("UID")).text = str(uid)
                ET.SubElement(t, E("ID")).text = str(uid)
                ET.SubElement(t, E("Name")).text = getattr(item, "name", "") or f"Task {uid}"
                ET.SubElement(t, E("Type")).text = "0"
                ET.SubElement(t, E("IsNull")).text = "0"
                ET.SubElement(t, E("CreateDate")).text = now

                ET.SubElement(t, E("WBS")).text = outline_number
                ET.SubElement(t, E("OutlineNumber")).text = outline_number
                ET.SubElement(t, E("OutlineLevel")).text = str(outline_level)

                ET.SubElement(t, E("Start")).text = start_iso
                ET.SubElement(t, E("Finish")).text = finish_iso
                ET.SubElement(t, E("Duration")).text = dur_str
                # 7 = Days, 5 = Hours in MSPDI; since we emit hours, set Hours (5)
                ET.SubElement(t, E("DurationFormat")).text = "5"

                # Flags
                ET.SubElement(t, E("Summary")).text = "1" if row_type == "Summary" else "0"
                ET.SubElement(t, E("Milestone")).text = "1" if row_type == "Milestone" else "0"
                ET.SubElement(t, E("Active")).text = "1"
                ET.SubElement(t, E("Manual")).text = "0"

                # Keep for predecessor linking pass
                flat_items.append((item, uid))

                # Recurse
                walk(getattr(item, "children", []), outline_level+1, prefix_numbers + [idx])

        # Kick off traversal
        try:
            root_items = list(self.template_items)
        except Exception:
            messagebox.showerror("Export error", "No items to export (self.template_items missing).")
            return

        walk(root_items, outline_level=1)

        # Add PredecessorLink per task
        # We expect attributes: predecessor_id, predecessor_type ('FS','SS','FF','SF'), and lag (days)
        TYPE_MAP = {"FS": "1", "SS": "2", "FF": "3", "SF": "4"}

        # Build mapping from UID -> <Task> element to attach links
        uid_to_task_el = { t.find(E("UID")).text: t for t in tasks_el.findall(E("Task")) }

        for item, uid in flat_items:
            pred_id = getattr(item, "predecessor_id", None)
            if not pred_id:
                continue
            pred_uid = id_map.get(pred_id)
            if not pred_uid:
                continue

            ptype = getattr(item, "predecessor_type", None) or "FS"
            ptype_val = TYPE_MAP.get(ptype.upper(), "1")
            lag_days = to_int(getattr(item, "lag", 0), 0)
            # MSPDI lag is in tenths of minutes if using numeric LinkLag; safer to use Duration+LagFormat:
            # We'll set LinkLag as hours*60*10, and LagFormat=5 (hours)
            link_lag_tenths_min = lag_days * HOURS_PER_DAY * 60 * 10

            task_el = uid_to_task_el.get(str(uid))
            if task_el is None:
                continue
            pl = ET.SubElement(task_el, E("PredecessorLink"))
            ET.SubElement(pl, E("PredecessorUID")).text = str(pred_uid)
            ET.SubElement(pl, E("Type")).text = ptype_val   # 1=FS,2=SS,3=FF,4=SF
            ET.SubElement(pl, E("CrossProject")).text = "0"
            ET.SubElement(pl, E("LinkLag")).text = str(link_lag_tenths_min)
            ET.SubElement(pl, E("LagFormat")).text = "5"    # 5 = Hours

        # Write file
        tree = ET.ElementTree(proj)
        tree.write(filename, encoding="utf-8", xml_declaration=True)

        messagebox.showinfo("Export complete", f"Exported MSPDI XML for ProjectLibre:\n{filename}\n\n"
                                               "Open ProjectLibre → File → Open → select this XML.")

    

    def export_to_smartsheet_csv(self):

        from datetime import date, datetime
        from tkinter import filedialog, messagebox

        # Try to ensure we have up-to-date calculations
        try:
            if hasattr(self, "calculate_all_dates") and callable(self.calculate_all_dates):
                self.calculate_all_dates()
        except Exception:
            pass  # continue with whatever is currently in memory

        # Ask where to save
        filename = filedialog.asksaveasfilename(
            title="Export for Smartsheet",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")]
        )
        if not filename:
            return

        headers = [
            "Task Name", "Start", "Finish", "Duration", "Predecessors",
            "ID", "Parent ID", "Price", "Type", "Indent"
        ]

        def _fmt_date(val):
            """Return YYYY-MM-DD or ''."""
            if not val:
                return ""
            if isinstance(val, datetime):
                return val.date().isoformat()
            if isinstance(val, date):
                return val.isoformat()
            s = str(val).strip()
            # ISO attempt
            try:
                return datetime.fromisoformat(s).date().isoformat()
            except Exception:
                # Common US format
                try:
                    return datetime.strptime(s, "%m/%d/%Y").date().isoformat()
                except Exception:
                    return s  # last resort; Smartsheet may still parse it

        def _coerce_int(x, default=0):
            try:
                return int(x)
            except Exception:
                return default

        def _price_val(it):
            try:
                return float(it.price) if it.price not in (None, "") else 0.0
            except Exception:
                return 0.0

        def _pred_string(it):
            """Build Smartsheet-friendly predecessor string 'IDType±Lag'."""
            pid = getattr(it, "predecessor_id", None)
            if not pid:
                return ""
            ptype = getattr(it, "predecessor_type", None) or "FS"
            lag = _coerce_int(getattr(it, "lag", 0), 0)
            lag_str = f"+{lag}d" if lag > 0 else (f"{lag}d" if lag < 0 else "")
            return f"{pid}{ptype}{lag_str}"

        def _classify(it):
            # Summary if it has children
            has_children = bool(getattr(it, "children", None))
            if has_children:
                return "Summary"
            # Milestone if flagged or duration coerces to 0
            if getattr(it, "is_milestone", False) or _coerce_int(getattr(it, "duration", 0), 0) == 0:
                return "Milestone"
            return "Task"

        # Fallback project start (string)
        proj_start = ""
        try:
            proj_start = str(self.project_start_date.get()).strip()
        except Exception:
            pass

        rows = []

        def _walk(items, parent_id=None, indent=0):
            for it in items or []:
                # Skip disabled rows
                enabled = True
                try:
                    enabled = bool(it.enabled.get())
                except Exception:
                    enabled = True
                if not enabled:
                    # Still descend, in case enabled children exist under a disabled header
                    _walk(getattr(it, "children", []), parent_id=getattr(it, "id", parent_id), indent=indent+1)
                    continue

                row_type = _classify(it)

                # Pull raw dates
                raw_start = getattr(it, "start_date", "") or ""
                raw_finish = getattr(it, "end_date", "") or ""
                duration = getattr(it, "duration", 0)

                # Normalize strings/ints
                start = str(raw_start).strip()
                finish = str(raw_finish).strip()
                dur_int = _coerce_int(duration, 0)

                if row_type == "Summary":
                    # Blank timing; Gantt rolls up from children once indented in Smartsheet
                    start_out = ""
                    finish_out = ""
                    pred_out = ""
                    dur_out = ""
                elif row_type == "Milestone":
                    # Force 0-day; if no date, fall back to project start if available
                    if not start and proj_start:
                        start = proj_start
                    if not start and finish:
                        start = finish
                    if not finish:
                        finish = start
                    start_out = _fmt_date(start) if start else ""
                    finish_out = _fmt_date(finish) if finish else start_out
                    dur_out = 0
                    pred_out = _pred_string(it)
                else:
                    # Task: fill missing with project start if available
                    if not start and proj_start:
                        start = proj_start
                    if not finish and start:
                        finish = start
                    start_out = _fmt_date(start) if start else ""
                    finish_out = _fmt_date(finish) if finish else ""
                    dur_out = dur_int
                    pred_out = _pred_string(it)

                rows.append({
                    "Task Name": getattr(it, "name", ""),
                    "Start": start_out,
                    "Finish": finish_out,
                    "Duration": dur_out,
                    "Predecessors": pred_out,
                    "ID": getattr(it, "id", ""),
                    "Parent ID": parent_id if parent_id is not None else "",
                    "Price": _price_val(it),
                    "Type": row_type,
                    "Indent": indent
                })

                # Recurse into children
                _walk(getattr(it, "children", []), parent_id=getattr(it, "id", ""), indent=indent+1)

        # Kick off traversal from your top-level list
        try:
            root_items = list(self.template_items)
        except Exception:
            messagebox.showerror("Export error", "No items to export (self.template_items missing).")
            return

        _walk(root_items, parent_id=None, indent=0)

        # Write CSV
        try:
            with open(filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
                writer.writeheader()
                for r in rows:
                    writer.writerow(r)
        except Exception as e:
            messagebox.showerror("Export error", f"Failed to write CSV:\n{e}")
            return

        messagebox.showinfo(
            "Export complete",
            f"Exported {len(rows)} rows to:\n{filename}\n\n"
            "Smartsheet tips:\n"
            "• Map 'Task Name' to Primary Column\n"
            "• Map Start / Finish (date), Duration (number), Predecessors\n"
            "• Keep 'Parent ID' & 'Indent' as helper columns\n"
            "• Indent rows in Smartsheet to make Summary roll-ups\n"
            "• Milestones render when Duration=0 and Start=Finish"
        )


    def create_template_structure(self):
        """Create the default template structure with sequential predecessors."""
        items = []
        all_tasks = []

        def create_and_collect_task(name, duration, price, is_milestone, indent_level):
            self.task_counter += 1
            item = ProposalItem(name, duration, price, "", is_milestone, indent_level, self.task_counter)
            if not is_milestone:
                all_tasks.append(item)
            return item

        # Project Initiation
        proj_init = create_and_collect_task("Project Initiation", 0, 0, True, 0)
        proj_init.children = [
            create_and_collect_task("Deposit & Contract Signed", 0, 0, False, 1),
            create_and_collect_task("Notice to Proceed", 0, 0, False, 1),
            create_and_collect_task("Civil Start - Civil Due Diligence", 1, 0, False, 1),
            create_and_collect_task("Electrical Start - Electrical Due Diligence", 1, 0, False, 1),
        ]
        items.append(proj_init)

        # Civil Engineering
        civil_eng = create_and_collect_task("Civil Engineering", 0, 0, True, 0)
        design_30_civil = create_and_collect_task("30% Design", 0, 0, True, 1)
        design_30_civil.children = [
            create_and_collect_task("30% - Planset/ Basis of Design", 20, 20000, False, 2),
            create_and_collect_task("Pre-Development Hydrology Study", 10, 10000, False, 2),
            create_and_collect_task("Client Review", 10, 0, False, 2),
        ]
        design_60_civil = create_and_collect_task("60% Design", 0, 0, True, 1)
        design_60_civil.children = [
            create_and_collect_task("60% - Planset", 25, 110000, False, 2),
            create_and_collect_task("Stormwater Pollution Prevention Plan", 10, 6000, False, 2),
            create_and_collect_task("Post-Development Hydrology Study", 10, 15000, False, 2),
            create_and_collect_task("Stormwater Management Report", 15, 12000, False, 2),
            create_and_collect_task("Client Review", 10, 0, False, 2),
        ]
        design_90_civil = create_and_collect_task("90% Design", 0, 0, True, 1)
        design_90_civil.children = [
            create_and_collect_task("90% - Planset", 5, 35000, False, 2),
            create_and_collect_task("Client Review", 10, 0, False, 2),
        ]
        ifc_design_civil = create_and_collect_task("IFC Design", 0, 0, True, 1)
        ifc_design_civil.children = [create_and_collect_task("IFC - Planset", 15, 56500, False, 2)]
        Studies_update = create_and_collect_task("Studies Updates", 5, 6500, True, 1)
        Studies_update.children = [
            create_and_collect_task("Stormwater Pollution Prevention Plan", 5, 1000, False, 2),
            create_and_collect_task("Post-Development Hydrology Study", 5, 2500, False, 2),
            create_and_collect_task("Stormwater Management Report", 5, 3000, False, 2),
        ]
        civil_eng.children = [design_30_civil, design_60_civil, design_90_civil, ifc_design_civil,Studies_update]
        items.append(civil_eng)

        # Electrical Engineering
        elec_eng = create_and_collect_task("Electrical Engineering", 0, 0, True, 0)
        design_30_elec = create_and_collect_task("30% Design", 0, 0, True, 1)
        design_30_elec.children = [
            create_and_collect_task("30% - Planset/Basis of Design", 11, 40000, False, 2),
            create_and_collect_task("Reactive Power Study", 6, 18500, False, 2),
            create_and_collect_task("MV - Short Circuit Study", 5, 6500, False, 2),
            create_and_collect_task("SAM Model", 3, 5000, False, 2),
            create_and_collect_task("PV SYST", 3, 5000, False, 2),
            create_and_collect_task("Client Review", 10, 0, False, 2),
        ]
        design_60_elec = create_and_collect_task("60% Design", 0, 0, True, 1)
        design_60_elec.children = [
            create_and_collect_task("60% - Planset", 14, 80000, False, 2),
            create_and_collect_task("DC - Short Circuit Study", 3, 6500, False, 2),
            create_and_collect_task("Under Ground Cable Thermal Study", 8, 10000, False, 2),
            create_and_collect_task("Grounding Study", 8, 13000, False, 2),
            create_and_collect_task("Client Review", 10, 0, False, 2),
        ]
        design_90_elec = create_and_collect_task("90% Design", 0, 0, True, 1)
        design_90_elec.children = [
            create_and_collect_task("90% - Planset", 13, 63500, False, 2),
            create_and_collect_task("Load Flow Study", 2, 13000, False, 2),
            create_and_collect_task("Coordination Study", 2, 9500, False, 2),
            create_and_collect_task("Arc Flash Study", 5, 13000, False, 2),
            create_and_collect_task("Client Review", 10, 0, False, 2),
        ]
        ifc_design_elec = create_and_collect_task("IFC Design", 0, 0, True, 1)
        ifc_design_elec.children = [
            create_and_collect_task("IFC - Planset", 10, 13000, False, 2),
        ]
        elec_eng.children = [design_30_elec, design_60_elec, design_90_elec, ifc_design_elec]
        items.append(elec_eng)

        # Structural Engineering
        struct_eng = create_and_collect_task("Structural Engineering", 0, 0, True, 0)
        struct_eng.children = [
            create_and_collect_task("Structural Engineering (Except racking foundation design)", 10, 25000, False, 1),
        ]
        items.append(struct_eng)

        # Project Closeout
        closeout = create_and_collect_task("Project Closeout", 0, 0, True, 0)
        items.append(closeout)

        # --- Set parent relationships ---
        def set_parents(item_list):
            for item in item_list:
                for child in item.children:
                    child.parent = item
                    if child.children:
                        set_parents(child.children)
        set_parents(items)
        
        # --- Set default sequential predecessors ---
        for i in range(1, len(all_tasks)):
            all_tasks[i].predecessor_id = all_tasks[i-1].id

        return items

    def setup_ui(self):
        """Sets up the main graphical user interface."""
        # Create main frames
        header_frame = ttk.Frame(self.root)
        header_frame.pack(fill=tk.X, padx=10, pady=5)

        top_button_frame = ttk.Frame(self.root)
        top_button_frame.pack(fill=tk.X, pady=5)

        content_frame = ttk.Frame(self.root)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        bottom_button_frame = ttk.Frame(self.root)
        bottom_button_frame.pack(fill=tk.X, padx=10, pady=10)


        # --- MODIFIED: Header Section with Title ---
        title_font = ("Arial", 16, "bold")
        title_label = ttk.Label(header_frame, text="Castillo Engineering Proposal Generator", font=title_font, anchor="center")
        title_label.pack(pady=(10, 5), fill=tk.X)

        info_container = ttk.Frame(header_frame)
        info_container.pack(pady=10)
        ttk.Label(info_container, text="Version:").grid(row=0, column=0, sticky=tk.E, padx=5)
        ttk.Entry(info_container, textvariable=self.version, width=15).grid(row=0, column=1, padx=5, sticky=tk.W)

        ttk.Label(info_container, text="Project Name:").grid(row=1, column=0, sticky=tk.E, padx=5)
        ttk.Entry(info_container, textvariable=self.project_name, width=40).grid(row=1, column=1, padx=5, sticky=tk.W)
        
        ttk.Label(info_container, text="Company Name:").grid(row=1, column=2, sticky=tk.E, padx=5)
        ttk.Entry(info_container, textvariable=self.company_name, width=40).grid(row=1, column=3, padx=5, sticky=tk.W)
        
        # --- MODIFIED: Added project start date with automatic recalculation ---
        
        ttk.Label(info_container, text="Project Start Date:").grid(row=2, column=0, sticky=tk.E, padx=5)
        if HAS_TKCAL:
            start_date_widget = DateEntry(
                info_container,
                textvariable=self.project_start_date,
                date_pattern="mm/dd/yy",
                width=15
            )
        else:
            # Fallback if tkcalendar isn't installed
            start_date_widget = ttk.Entry(info_container, textvariable=self.project_start_date, width=15)
        start_date_widget.grid(row=2, column=1, padx=5, sticky=tk.W)
        # --- MODIFICATION: Add trace for automatic recalculation ---
        self.project_start_date.trace_add("write", self.handle_project_start_change)

        
        ttk.Label(info_container, text="Company Logo:").grid(row=3, column=0, sticky=tk.E, padx=5, pady=(5,0))
        logo_entry = ttk.Entry(info_container, textvariable=self.logo_path, width=50, state='readonly')
        logo_entry.grid(row=3, column=1, columnspan=2, padx=5, pady=(5,0), sticky=tk.W)
        ttk.Button(info_container, text="Change Logo", command=self.change_logo).grid(row=3, column=3, padx=5, pady=(5,0), sticky=tk.W)
        
        # --- MODIFIED: Added client logo upload ---
        ttk.Label(info_container, text="Client Logo:").grid(row=4, column=0, sticky=tk.E, padx=5, pady=(5,0))
        client_logo_entry = ttk.Entry(info_container, textvariable=self.client_logo_path, width=50, state='readonly')
        client_logo_entry.grid(row=4, column=1, columnspan=2, padx=5, pady=(5,0), sticky=tk.W)
        ttk.Button(info_container, text="Change Logo", command=self.change_client_logo).grid(row=4, column=3, padx=5, pady=(5,0), sticky=tk.W)
        
        # --- MODIFIED: Added Gantt chart checkbox ---
        gantt_check = ttk.Checkbutton(info_container, text="Include Gantt Chart", variable=self.include_gantt)
        gantt_check.grid(row=5, column=1, columnspan=2, pady=(10,0), sticky=tk.W)


        # --- Top Button Container ---
        top_button_container = ttk.Frame(top_button_frame)
        top_button_container.pack()
        ttk.Button(top_button_container, text="Reset Predecessors", command=self.reset_predecessors).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_button_container, text="Clear Predecessors", command=self.clear_all_predecessors).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_button_container, text="Clear Prices", command=self.clear_all_prices).pack(side=tk.LEFT, padx=5)
        # --- MODIFICATION: Changed button text for Excel ---
        ttk.Button(top_button_container, text="Load Excel Template", command=self.load_template_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_button_container, text="Save Excel Template", command=self.save_template_excel).pack(side=tk.LEFT, padx=5)

        # --- Bottom Button Container ---
        ttk.Button(bottom_button_frame, text="Generate PDF", command=self.generate_pdf).pack(side=tk.RIGHT, padx=5)
        ttk.Button(bottom_button_frame, text="Export to Smartsheet", command=self.export_to_projectlibre_xml).pack(side=tk.RIGHT, padx=5)

        # --- MODIFICATION: Make "Calculate Dates" button unpin all dates ---
        ttk.Button(bottom_button_frame, text="Calculate Dates", command=lambda: self.calculate_all_dates(unpin_all=True)).pack(side=tk.LEFT, padx=5)
        ttk.Button(bottom_button_frame, text="Delete Item", command=self.delete_item).pack(side=tk.LEFT, padx=5)
        ttk.Button(bottom_button_frame, text="Add Custom Item", command=self.add_custom_item).pack(side=tk.LEFT, padx=5)
        

        # Content section with treeview
        tree_frame = ttk.Frame(content_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        column_ids = ('Predecessor', 'Type', 'Enabled', 'Duration', 'Price', 'Start Date', 'End Date')
        self.tree = ttk.Treeview(tree_frame, columns=column_ids, displaycolumns=column_ids, show='tree headings')
        
        self.tree.heading('#0', text='Task Name') # The tree column is '#0'
        self.tree.heading('Predecessor', text='Predecessor')
        self.tree.heading('Type', text='Type')
        self.tree.heading('Enabled', text='Enabled')
        self.tree.heading('Duration', text='Duration (days)')
        self.tree.heading('Price', text='Price ($)')
        self.tree.heading('Start Date', text='Start Date')
        self.tree.heading('End Date', text='End Date')

        self.tree.column('#0', width=250)
        self.tree.column('Predecessor', width=150)
        self.tree.column('Type', width=60, anchor='center')
        self.tree.column('Enabled', width=60, anchor='center')
        self.tree.column('Duration', width=100)
        self.tree.column('Price', width=100)
        self.tree.column('Start Date', width=100)
        self.tree.column('End Date', width=100)
        
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        self.tree.tag_configure('milestone', background='#E8E8E8', font=('Arial', 9, 'bold'))
        self.tree.tag_configure('task', font=('Arial', 9))
        self.tree.tag_configure('predecessor_highlight', background='lightgreen')
        self.tree.tag_configure('successor_highlight', background='lightcoral')
        self.tree.tag_configure('linking_highlight', background='lightblue')
        
        # Event Bindings
        self.tree.bind('<Double-1>', self.on_item_double_click)
        self.tree.bind('<Button-1>', self.on_item_click, add='+')
        self.tree.bind('<ButtonPress-1>', self.on_drag_start, add='+')
        self.tree.bind('<B1-Motion>', self.on_drag_motion, add='+')
        self.tree.bind('<ButtonRelease-1>', self.on_drag_release, add='+')
        self.tree.bind('<Control-ButtonPress-1>', self.on_link_start)
        self.tree.bind('<Control-B1-Motion>', self.on_link_drag)
        self.tree.bind('<Control-ButtonRelease-1>', self.on_link_drop)

    def handle_project_start_change(self, *args):
        """Handle changes to the main project start date."""
        new_date = self.project_start_date.get()
        try:
            # Validate the new date format first
            datetime.strptime(new_date, "%m/%d/%y")
        except ValueError:
            # If format is invalid during typing, just wait for a valid one
            return

        # Check if any tasks have pinned start dates
        has_pinned_dates = any(item.is_start_pinned for item in self.item_id_map.values())
        
        if has_pinned_dates:
            if messagebox.askyesno("Confirm Date Change", 
                                   "This will reset all manually entered start dates and recalculate the schedule from the new project start date. Are you sure you want to continue?"):
                self.last_project_start_date = new_date
                self.calculate_all_dates(unpin_all=True)
            else:
                # User cancelled, so revert the change in the Entry widget
                self.project_start_date.set(self.last_project_start_date)
        else:
            # No pinned dates, so just recalculate
            self.last_project_start_date = new_date
            self.calculate_all_dates()

    def clear_all_prices(self):
        """Sets the price of all items to zero."""
        if messagebox.askyesno("Confirm Clear", "Are you sure you want to clear all prices?"):
            for item in self.item_id_map.values():
                item.price = 0
            self.populate_tree()
            self.expand_all_items()

    def clear_all_predecessors(self):
        """Removes all predecessor links from all tasks."""
        if messagebox.askyesno("Confirm Clear", "Are you sure you want to clear all predecessor links?"):
            for item in self.item_id_map.values():
                item.predecessor_id = None
            self.populate_tree()
            self.expand_all_items()
    
    def reset_predecessors(self):
        """Resets all tasks to have a sequential predecessor link."""
        if messagebox.askyesno("Confirm Reset", "Are you sure you want to reset all predecessors to the default sequential order?"):
            ordered_tasks = []
            def get_tasks_in_order(tree_item_ids):
                for tree_id in tree_item_ids:
                    task_obj = self.tree_item_map.get(tree_id)
                    if task_obj and not task_obj.is_milestone and task_obj.enabled.get():
                        ordered_tasks.append(task_obj)
                    children = self.tree.get_children(tree_id)
                    if children:
                        get_tasks_in_order(children)

            get_tasks_in_order(self.tree.get_children())
            
            # First, clear all existing predecessors for non-milestone tasks
            for task in self.item_id_map.values():
                if not task.is_milestone:
                    task.predecessor_id = None

            if ordered_tasks:
                for i in range(1, len(ordered_tasks)):
                    ordered_tasks[i].predecessor_id = ordered_tasks[i-1].id
            
            self.calculate_all_dates() # Recalculate dates after resetting
            self.populate_tree()
            self.expand_all_items()

    def change_logo(self):
        """Open a file dialog to select a new logo file."""
        filepath = filedialog.askopenfilename(
            title="Select Logo File",
            filetypes=[("Image Files", "*.png *.jpg *.jpeg"), ("All files", "*.*")]
        )
        if filepath:
            self.logo_path.set(filepath)

    def change_client_logo(self):
        """Open a file dialog to select a new client logo file."""
        filepath = filedialog.askopenfilename(
            title="Select Client Logo File",
            filetypes=[("Image Files", "*.png *.jpg *.jpeg"), ("All files", "*.*")]
        )
        if filepath:
            self.client_logo_path.set(filepath)

    def on_drag_start(self, event):
        """Prepares for reordering an item or a column."""
        region = self.tree.identify("region", event.x, event.y)

        if region == "heading":
            self.column_drag_data["start_x"] = event.x
            self.column_drag_data["col_id"] = self.tree.identify_column(event.x)
        else:
            if self.tree.identify_column(event.x) != '#0': return 
            item_id = self.tree.identify_row(event.y)
            if item_id and self.tree.parent(item_id):
                self.drag_data["item"] = item_id
                self.drag_data["index"] = self.tree.index(item_id)

    def on_drag_motion(self, event):
        """Moves the dragged item or provides visual feedback for column drag."""
        if self.column_drag_data.get("col_id"):
            self.root.config(cursor="sb_h_double_arrow")
            return

        if not self.drag_data.get("item"): return
        drag_item = self.drag_data["item"]
        parent = self.tree.parent(drag_item)
        self.tree.move(drag_item, parent, self.tree.index(self.tree.identify_row(event.y)))

    def on_drag_release(self, event):
        """Finalizes the item's new position or reorders columns."""
        if self.column_drag_data.get("col_id"):
            self.root.config(cursor="")
            dragged_col_id = self.column_drag_data["col_id"]
            target_col_id = self.tree.identify_column(event.x)
            
            if dragged_col_id and target_col_id:
                dragged_index = int(dragged_col_id.replace('#','')) - 1
                target_index = int(target_col_id.replace('#','')) - 1
                
                cols = list(self.tree['displaycolumns'])
                cols.insert(target_index, cols.pop(dragged_index))
                self.tree['displaycolumns'] = tuple(cols)
            self.column_drag_data = {}
            return

        if not self.drag_data.get("item"): return
        dragged_id = self.drag_data["item"]
        parent_id = self.tree.parent(dragged_id)
        if not parent_id: 
            self.drag_data = {"item": None, "index": 0}
            return
        new_index = self.tree.index(dragged_id)
        dragged_item_obj = self.tree_item_map[dragged_id]
        parent_item_obj = self.tree_item_map.get(parent_id)
        if parent_item_obj and dragged_item_obj in parent_item_obj.children:
            parent_item_obj.children.remove(dragged_item_obj)
            parent_item_obj.children.insert(new_index, dragged_item_obj)
        self.drag_data = {"item": None, "index": 0}
        
    def _toggle_children_enabled(self, item, enabled):
        """Recursively sets the enabled state for an item and all its children."""
        item.enabled.set(enabled)
        for child in item.children:
            self._toggle_children_enabled(child, enabled)

    def on_item_click(self, event):
        """Handle single click for checkbox toggle, opening type dropdown, and highlighting."""
        item_id = self.tree.identify_row(event.y)

        if not item_id:
            self.clear_highlights()
            return

        region = self.tree.identify("region", event.x, event.y)
        if region != "cell": 
            self.highlight_dependencies(item_id)
            return

        column_id = self.tree.identify_column(event.x)
        item = self.tree_item_map.get(item_id)
        if not item: return

        # Get current column display order to find correct index
        display_cols = list(self.tree['displaycolumns'])
        
        if column_id == f"#{display_cols.index('Enabled') + 1}":
            new_state = not item.enabled.get()
            self._toggle_children_enabled(item, new_state)
            self.populate_tree() # Refresh to show visual updates for all children
            self.expand_all_items()
            self.calculate_all_dates() # Recalculate after state change
        
        elif column_id == f"#{display_cols.index('Type') + 1}" and not item.is_milestone:
            if item.predecessor_id:
                self.edit_type_cell(item_id, item, column_id)
        
        self.highlight_dependencies(item_id)

    def on_link_start(self, event):
        """Starts a predecessor link drag operation."""
        item_id = self.tree.identify_row(event.y)
        item = self.tree_item_map.get(item_id)
        if item and not item.is_milestone:
            self.link_drag_data["start_item_id"] = item_id

    def on_link_drag(self, event):
        """Updates the visual highlight while dragging."""
        if not self.link_drag_data.get("start_item_id"):
            return

        last_hover_id = self.link_drag_data.get("last_hover_id")
        if last_hover_id and self.tree.exists(last_hover_id):
            tags = list(self.tree.item(last_hover_id, 'tags'))
            if 'linking_highlight' in tags:
                tags.remove('linking_highlight')
                self.tree.item(last_hover_id, tags=tuple(tags))
        
        current_hover_id = self.tree.identify_row(event.y)
        start_item_id = self.link_drag_data["start_item_id"]
        
        if current_hover_id and current_hover_id != start_item_id:
            item = self.tree_item_map.get(current_hover_id)
            if item and not item.is_milestone:
                tags = list(self.tree.item(current_hover_id, 'tags'))
                if 'linking_highlight' not in tags:
                    tags.append('linking_highlight')
                    self.tree.item(current_hover_id, tags=tuple(tags))
                self.link_drag_data["last_hover_id"] = current_hover_id
            else:
                self.link_drag_data["last_hover_id"] = None
        else:
            self.link_drag_data["last_hover_id"] = None

    def on_link_drop(self, event):
        """Finalizes the predecessor link."""
        start_item_id = self.link_drag_data.get("start_item_id")
        if not start_item_id:
            return

        last_hover_id = self.link_drag_data.get("last_hover_id")
        if last_hover_id and self.tree.exists(last_hover_id):
            tags = list(self.tree.item(last_hover_id, 'tags'))
            if 'linking_highlight' in tags:
                tags.remove('linking_highlight')
                self.tree.item(last_hover_id, tags=tuple(tags))

        end_item_id = self.tree.identify_row(event.y)
        start_item = self.tree_item_map.get(start_item_id)
        end_item = self.tree_item_map.get(end_item_id)

        if (start_item and end_item and
            not start_item.is_milestone and not end_item.is_milestone and
            start_item.id != end_item.id):
            
            start_item.predecessor_id = end_item.id
            start_item.predecessor_type = 'FS'
            start_item.lag = 0
            self.update_item_display(start_item_id, start_item)
            self.highlight_dependencies(start_item_id)

        self.link_drag_data = {"start_item_id": None, "last_hover_id": None}

    def clear_highlights(self):
        """Removes all dependency highlighting from the tree."""
        for item_id in self.tree_item_map:
            if self.tree.exists(item_id):
                current_tags = list(self.tree.item(item_id, 'tags'))
                if 'predecessor_highlight' in current_tags:
                    current_tags.remove('predecessor_highlight')
                if 'successor_highlight' in current_tags:
                    current_tags.remove('successor_highlight')
                self.tree.item(item_id, tags=tuple(current_tags))

    def highlight_dependencies(self, selected_item_id):
        """Highlights the predecessor and successors of the selected item."""
        self.clear_highlights()
        
        selected_item = self.tree_item_map.get(selected_item_id)
        if not selected_item or selected_item.is_milestone:
            return

        if selected_item.predecessor_id:
            for tree_id, item_obj in self.tree_item_map.items():
                if item_obj.id == selected_item.predecessor_id:
                    if self.tree.exists(tree_id):
                        current_tags = list(self.tree.item(tree_id, 'tags'))
                        if 'predecessor_highlight' not in current_tags:
                            current_tags.append('predecessor_highlight')
                        self.tree.item(tree_id, tags=tuple(current_tags))
                    break
        
        for tree_id, item_obj in self.tree_item_map.items():
            if item_obj.predecessor_id == selected_item.id:
                if self.tree.exists(tree_id):
                    current_tags = list(self.tree.item(tree_id, 'tags'))
                    if 'successor_highlight' not in current_tags:
                        current_tags.append('successor_highlight')
                    self.tree.item(tree_id, tags=tuple(current_tags))

    def populate_tree(self):
        """Populate the treeview with template items and build ID maps."""
        expanded_items = []
        for item_id in self.tree.get_children():
            if self.tree.item(item_id, 'open'):
                expanded_items.append(self.tree.item(item_id, 'text'))
            expanded_items.extend(self.get_expanded_children(item_id))
        
        self.tree.delete(*self.tree.get_children())
        self.tree_item_map = {}
        self.item_id_map = {}
        
        def build_id_map(items):
            for item in items:
                self.item_id_map[item.id] = item
                if item.children:
                    build_id_map(item.children)
        build_id_map(self.template_items)
        
        for item in self.template_items:
            item_id = self.add_item_to_tree(item, '')
            if any(expanded in self.tree.item(item_id, 'text') for expanded in expanded_items):
                self.tree.item(item_id, open=True)
    
    def get_expanded_children(self, item_id):
        """Get all expanded children recursively."""
        expanded = []
        for child_id in self.tree.get_children(item_id):
            if self.tree.item(child_id, 'open'):
                expanded.append(self.tree.item(child_id, 'text'))
                expanded.extend(self.get_expanded_children(child_id))
        return expanded
    
    def expand_all_items(self):
        """Expand all items in the tree by default."""
        def expand_children(item_id):
            self.tree.item(item_id, open=True)
            for child_id in self.tree.get_children(item_id):
                expand_children(child_id)
        for item_id in self.tree.get_children():
            expand_children(item_id)
    
    def add_item_to_tree(self, item, parent_id):
        """Recursively add items to treeview."""
        # Update display name to include the unique ID
        display_name = f"{'  ' * item.indent_level}({item.id}) {item.name}"
        enabled_text = "✓" if item.enabled.get() else "✗"
        
        predecessor_text = ""
        predecessor_type_text = ""
        if item.predecessor_id and item.predecessor_id in self.item_id_map:
            pred_item = self.item_id_map[item.predecessor_id]
            lag_str = f" +{item.lag}d" if item.lag > 0 else f" {item.lag}d" if item.lag < 0 else ""
            predecessor_text = f"({pred_item.id}) {pred_item.name[:15]}{lag_str}"
            predecessor_type_text = item.predecessor_type
            
        item_id = self.tree.insert(parent_id, 'end', text=display_name, 
                                           values=(predecessor_text, predecessor_type_text, enabled_text, item.duration, f"${item.price:,}", item.start_date, item.end_date),
                                           tags=('milestone' if item.is_milestone else 'task',))
        self.tree_item_map[item_id] = item
        for child in item.children:
            self.add_item_to_tree(child, item_id)
        return item_id

    def on_item_double_click(self, event):
        """Handle double-click for inline editing or opening predecessor dialog."""
        item_id = self.tree.identify_row(event.y)
        column_id = self.tree.identify_column(event.x)
        if not item_id or not column_id: return
        item = self.tree_item_map.get(item_id)
        if not item: return

        display_cols = list(self.tree['displaycolumns'])
        
        # New condition to edit task name in the first column
        if column_id == '#0':
            self.edit_cell(item_id, item, 'name', column_id)
        elif column_id == f"#{display_cols.index('Enabled') + 1}":
            new_state = not item.enabled.get()
            self._toggle_children_enabled(item, new_state)
            self.populate_tree()
            self.expand_all_items()
            self.calculate_all_dates()
        elif not item.is_milestone:
            if column_id == f"#{display_cols.index('Duration') + 1}": self.edit_cell(item_id, item, 'duration', column_id)
            elif column_id == f"#{display_cols.index('Price') + 1}": self.edit_cell(item_id, item, 'price', column_id)
            elif column_id == f"#{display_cols.index('Predecessor') + 1}": self.edit_predecessor(item_id)
            elif column_id == f"#{display_cols.index('Start Date') + 1}": self.edit_cell(item_id, item, 'start_date', column_id)
            elif column_id == f"#{display_cols.index('Type') + 1}":  self.edit_type_cell(item_id, item, column_id)

    def edit_cell(self, item_id, item, attribute, column_id):
        """Create inline editor for a cell."""
        if self.current_editor: self.current_editor.destroy()
        bbox = self.tree.bbox(item_id, column_id)
        if not bbox: return
        x, y, w, h = bbox
        
        current_value = getattr(item, attribute)
        entry = tk.Entry(self.tree, font=('Arial', 9))
        self.current_editor = entry
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, str(current_value))
        entry.select_range(0, tk.END)
        entry.focus()
        
        def save_edit(event=None):
            try:
                new_value = entry.get()
                if attribute == 'duration':
                    setattr(item, attribute, math.ceil(float(new_value)) if new_value else 0)
                elif attribute == 'price':
                    setattr(item, attribute, int(new_value.replace('$', '').replace(',', '')) if new_value else 0)
                elif attribute == 'start_date':
                    if new_value:
                        datetime.strptime(new_value, "%m/%d/%y") # Validate date format
                        item.start_date = new_value
                        item.is_start_pinned = True
                    else:
                        item.start_date = ""
                        item.is_start_pinned = False
                    self.calculate_all_dates()
                    # No need to call update_item_display, calculate_all_dates will refresh the tree
                    if entry and entry.winfo_exists(): entry.destroy()
                    self.current_editor = None
                    return
                elif attribute == 'name':
                    if new_value.strip(): # Don't allow empty names
                        setattr(item, attribute, new_value)
                        self.populate_tree() # Refresh the entire tree to update predecessors
                        self.expand_all_items()
                        if entry and entry.winfo_exists(): entry.destroy()
                        self.current_editor = None
                        return
                    else:
                         messagebox.showerror("Invalid Name", "Task name cannot be empty.")
                         return
                else:
                    setattr(item, attribute, new_value)
                
                self.update_item_display(item_id, item)

            except ValueError:
                if attribute == 'start_date':
                    messagebox.showerror("Invalid Date", "Please use MM/DD/YY format.")
                else:
                    messagebox.showerror("Invalid Input", f"Please enter a valid number for {attribute}.")
            except (tk.TclError):
                pass # Ignore Tcl errors which can happen on widget destruction
            finally:
                if entry and entry.winfo_exists(): entry.destroy()
                self.current_editor = None
        
        entry.bind('<Return>', save_edit)
        entry.bind('<KP_Enter>', save_edit)
        entry.bind('<Escape>', lambda e: entry.destroy())
        entry.bind('<FocusOut>', save_edit)
    def edit_type_cell(self, item_id, item, column_id):
        """Inline editor for the 'Type' (FS/SS/FF/SF) column."""
        # Must have a predecessor to set a type
        if not getattr(item, "predecessor_id", None):
            messagebox.showinfo("No predecessor", "Set a predecessor before changing the type.")
            return

        # If another editor is open, close it
        if self.current_editor:
            try:
                self.current_editor.destroy()
            except Exception:
                pass
            self.current_editor = None

        bbox = self.tree.bbox(item_id, column_id)
        if not bbox:
            return
        x, y, w, h = bbox

        # Create a combobox in-place
        type_var = tk.StringVar(value=(item.predecessor_type or "FS"))
        type_combo = ttk.Combobox(
            self.tree,
            textvariable=type_var,
            values=["FS", "SS", "FF", "SF"],
            state="readonly",
            width=max(4, int(w/8))  # rough fit
        )
        self.current_editor = type_combo
        type_combo.place(x=x, y=y, width=w, height=h)
        type_combo.focus()
        type_combo.dropdown_visible = True  # hint for some ttk themes

        def commit_and_close(*_):
            sel = type_var.get().strip().upper() or "FS"
            if sel not in ("FS", "SS", "FF", "SF"):
                sel = "FS"
            item.predecessor_type = sel
            # Recalculate and refresh visuals
            try:
                self.calculate_all_dates()
            except Exception:
                # As a fallback, at least refresh the one row
                self.update_item_display(item_id, item)
            finally:
                if type_combo and type_combo.winfo_exists():
                    type_combo.destroy()
                self.current_editor = None

        # Save on selection or focus-out; Escape cancels
        type_combo.bind("<<ComboboxSelected>>", commit_and_close)
        type_combo.bind("<FocusOut>", commit_and_close)
        type_combo.bind("<Return>", commit_and_close)
        type_combo.bind("<Escape>", lambda e: (type_combo.destroy(), setattr(self, "current_editor", None)))

    def get_item_path(self, item):
        """Build the full path for a given item, handling the base case."""
        path = []
        current_item = item
        while current_item:
            path.append(f"({current_item.id}) {current_item.name}")
            current_item = current_item.parent
        return " > ".join(reversed(path))

    def edit_predecessor(self, item_id):
        """Open a dialog to set an item's predecessor."""
        item_to_edit = self.tree_item_map.get(item_id)
        if not item_to_edit: return

        dialog = tk.Toplevel(self.root)
        dialog.title(f"Set Predecessor for '{item_to_edit.name}'")
        dialog.geometry("550x200") # Increased width for longer names
        dialog.transient(self.root)
        dialog.grab_set()

        # --- MODIFICATION: Use full path for unique predecessor names ---
        possible_preds = {self.get_item_path(i): i.id for i in self.item_id_map.values() if not i.is_milestone and i.id != item_to_edit.id}
        
        frame = ttk.Frame(dialog, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="Predecessor Task:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        pred_var = tk.StringVar()
        pred_combo = ttk.Combobox(frame, textvariable=pred_var, values=list(possible_preds.keys()), width=60)
        pred_combo.grid(row=0, column=1, columnspan=2, padx=5, pady=5)
        
        ttk.Label(frame, text="Lag (days):").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        lag_var = tk.IntVar(value=item_to_edit.lag)
        lag_entry = ttk.Entry(frame, textvariable=lag_var, width=8)
        lag_entry.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)

        if item_to_edit.predecessor_id:
            for path, p_id in possible_preds.items():
                if p_id == item_to_edit.predecessor_id:
                    pred_combo.set(path)
                    break
        
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=2, column=0, columnspan=3, pady=20)

        def save_predecessor():
            selected_path = pred_var.get()
            if selected_path in possible_preds:
                item_to_edit.predecessor_id = possible_preds[selected_path]
                item_to_edit.lag = lag_var.get()
            self.calculate_all_dates()
            self.highlight_dependencies(item_id)
            dialog.destroy()
        
        def clear_predecessor():
            item_to_edit.predecessor_id = None
            item_to_edit.is_start_pinned = False # Unpin if predecessor is removed
            self.calculate_all_dates()
            self.highlight_dependencies(item_id)
            dialog.destroy()

        ttk.Button(button_frame, text="Save", command=save_predecessor).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="Clear", command=clear_predecessor).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=10)
    
    def update_item_display(self, item_id, item):
        """Update a single item's display without refreshing entire tree."""
        enabled_text = "✓" if item.enabled.get() else "✗"
        predecessor_text = ""
        predecessor_type_text = ""
        if item.predecessor_id and item.predecessor_id in self.item_id_map:
            pred_item = self.item_id_map[item.predecessor_id]
            lag_str = f" +{item.lag}d" if item.lag > 0 else f" {item.lag}d" if item.lag < 0 else ""
            predecessor_text = f"({pred_item.id}) {pred_item.name[:15]}{lag_str}"
            predecessor_type_text = item.predecessor_type
        
        value_map = {
            'Predecessor': predecessor_text,
            'Type': predecessor_type_text,
            'Enabled': enabled_text,
            'Duration': item.duration,
            'Price': f"${item.price:,}",
            'Start Date': item.start_date,
            'End Date': item.end_date
        }
        
        display_cols = self.tree['displaycolumns']
        values_tuple = tuple(value_map[col_id] for col_id in display_cols)

        if self.tree.exists(item_id):
            self.tree.item(item_id, values=values_tuple)
    
    def add_custom_item(self):
        """Add a custom item to the project."""
        selection = self.tree.selection()
        parent_item = self.tree_item_map.get(selection[0]) if selection else None
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Add Custom Item")
        dialog.geometry("400x280") # Increased height for new checkbox
        
        ttk.Label(dialog, text="Name:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        name_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=name_var, width=40).grid(row=0, column=1, padx=5, pady=5)
        ttk.Label(dialog, text="Duration (days):").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        duration_var = tk.DoubleVar()
        ttk.Entry(dialog, textvariable=duration_var, width=20).grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)
        ttk.Label(dialog, text="Price ($):").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        price_var = tk.IntVar()
        ttk.Entry(dialog, textvariable=price_var, width=20).grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)
        ttk.Label(dialog, text="Is Milestone:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=5)
        is_milestone_var = tk.BooleanVar()
        ttk.Checkbutton(dialog, variable=is_milestone_var).grid(row=3, column=1, sticky=tk.W, padx=5, pady=5)

        # --- MODIFICATION: Add "New Section" checkbox ---
        is_new_section_var = tk.BooleanVar()
        ttk.Checkbutton(dialog, text="Add as new section", variable=is_new_section_var).grid(row=4, column=1, sticky=tk.W, padx=5, pady=5)
        
        button_frame = ttk.Frame(dialog)
        button_frame.grid(row=5, column=0, columnspan=2, pady=20)
        
        def add_item():
            is_new_section = is_new_section_var.get()
            
            # --- MODIFICATION: Round duration up ---
            duration = math.ceil(duration_var.get())
            self.task_counter += 1
            new_id = self.task_counter

            if is_new_section:
                # Add as a new top-level section
                new_item = ProposalItem(name_var.get(), duration, price_var.get(), "", True, 0, new_id)
                self.template_items.append(new_item)
            else:
                # Add as a child of the selected item (or as a top-level item if nothing is selected)
                indent_level = parent_item.indent_level + 1 if parent_item else 0
                new_item = ProposalItem(name_var.get(), duration, price_var.get(), "", is_milestone_var.get(), indent_level, new_id)
                if parent_item:
                    new_item.parent = parent_item
                    parent_item.children.append(new_item)
                else:
                    self.template_items.append(new_item)

            self.populate_tree()
            self.expand_all_items()
            dialog.destroy()
        
        ttk.Button(button_frame, text="Add", command=add_item).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
    
    def delete_item(self):
        """Delete selected item."""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select an item to delete.")
            return
        item_id = selection[0]
        item = self.tree_item_map.get(item_id)
        if not item: return
        
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete '{item.name}' and all its children?"):
            if item.parent: item.parent.children.remove(item)
            else: self.template_items.remove(item)
            self.populate_tree()
            self.expand_all_items()
    
    def _add_business_days(self, start_date_str, days_to_add):
        """Adds or subtracts business days from a given date string."""
        if not start_date_str: return ""
        try:
            current_date = datetime.strptime(start_date_str, "%m/%d/%y")
        except ValueError:
            return ""

        step = timedelta(days=1) if days_to_add >= 0 else timedelta(days=-1)
        days_counted = 0
        
        days_to_add = int(days_to_add)

        if days_to_add > 0: days_to_add -= 1

        while current_date.weekday() >= 5 or current_date in self.us_holidays:
            current_date += timedelta(days=1)

        while days_counted < abs(days_to_add):
            current_date += step
            if current_date.weekday() < 5 and current_date not in self.us_holidays:
                days_counted += 1

        while current_date.weekday() >= 5 or current_date in self.us_holidays:
            current_date += step

        return current_date.strftime("%m/%d/%y")

    def _get_business_days_between(self, start_date_str, end_date_str):
        """Calculate the number of business days between two dates."""
        try:
            start_date = datetime.strptime(start_date_str, "%m/%d/%y")
            end_date = datetime.strptime(end_date_str, "%m/%d/%y")
            
            if start_date > end_date:
                return 0
            
            # Count business days
            days = (end_date - start_date).days + 1
            business_days = 0
            for i in range(days):
                day = start_date + timedelta(days=i)
                if day.weekday() < 5 and day not in self.us_holidays:
                    business_days += 1
            
            return business_days
        except (ValueError, TypeError):
            return 0

    def calculate_all_dates(self, unpin_all=False):
        """
        Calculate all dates based on dependencies and durations.
        If unpin_all is True, it will ignore all manually set dates.
        """
        if unpin_all:
            for item in self.item_id_map.values():
                item.is_start_pinned = False

        for item in self.item_id_map.values():
            if not item.is_start_pinned:
                item.start_date = ""
                item.end_date = ""

        all_tasks = [item for item in self.item_id_map.values() if item.enabled.get() and not item.is_milestone]

        graph = {item.id: [] for item in all_tasks}
        in_degree = {item.id: 0 for item in all_tasks}
        for item in all_tasks:
            if item.predecessor_id and item.predecessor_id in self.item_id_map:
                pred = self.item_id_map[item.predecessor_id]
                if pred.enabled.get():
                    graph[item.predecessor_id].append(item.id)
                    in_degree[item.id] += 1

        queue = [item_id for item_id in in_degree if in_degree[item_id] == 0]
        sorted_order = []
        while queue:
            u_id = queue.pop(0)
            sorted_order.append(u_id)
            for v_id in graph.get(u_id, []):
                in_degree[v_id] -= 1
                if in_degree[v_id] == 0:
                    queue.append(v_id)

        if len(sorted_order) != len(all_tasks):
            messagebox.showerror("Calculation Error", "A circular dependency was detected. Please fix the predecessors.")
            return

        project_start = self.project_start_date.get()
        for item_id in sorted_order:
            item = self.item_id_map[item_id]

            if not item.is_start_pinned:
                pred_item = self.item_id_map.get(item.predecessor_id) if item.predecessor_id else None
                
                if pred_item and pred_item.enabled.get() and pred_item.end_date:
                    if item.predecessor_type == 'FS':
                        item.start_date = self._add_business_days(pred_item.end_date, item.lag + 1)
                    elif item.predecessor_type == 'SS':
                        item.start_date = self._add_business_days(pred_item.start_date, item.lag)
                    elif item.predecessor_type == 'FF':
                        finish_date = self._add_business_days(pred_item.end_date, item.lag)
                        item.start_date = self._add_business_days(finish_date, -item.duration + 1)
                    elif item.predecessor_type == 'SF':
                        finish_date = self._add_business_days(pred_item.start_date, item.lag)
                        item.start_date = self._add_business_days(finish_date, -item.duration + 1)
                else:
                    item.start_date = project_start
            
            item.end_date = self._add_business_days(item.start_date, item.duration)

        def calculate_milestone_rollup(items):
            for item in items:
                if item.enabled.get() and item.is_milestone and item.children:
                    calculate_milestone_rollup(item.children)
                    enabled_children = [c for c in item.children if c.enabled.get()]
                    if enabled_children:
                        valid_starts = [datetime.strptime(c.start_date, "%m/%d/%y") for c in enabled_children if c.start_date]
                        valid_ends = [datetime.strptime(c.end_date, "%m/%d/%y") for c in enabled_children if c.end_date]
                        
                        if valid_starts: item.start_date = min(valid_starts).strftime("%m/%d/%y")
                        if valid_ends: item.end_date = max(valid_ends).strftime("%m/%d/%y")
                        
                        # --- MODIFICATION: Correct milestone duration calculation ---
                        item.duration = self._get_business_days_between(item.start_date, item.end_date)
                        item.price = sum(c.price for c in enabled_children)
        calculate_milestone_rollup(self.template_items)
        end_date = self.get_project_end_date()
        if end_date:
            print(f"Project End Date: {end_date}")
        self.populate_tree()
        self.expand_all_items()

    def generate_pdf(self):
        """Generate a single PDF proposal with the Gantt chart on page 2."""
        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            title="Save Proposal As"
        )
        if not filename: return
        
        try:
            self.create_pdf(filename)
            messagebox.showinfo("Success", f"Successfully generated proposal:\n{os.path.basename(filename)}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate PDF: {str(e)}")

    def _setup_reportlab_styles(self, num_rows):
        """Dynamically create ReportLab styles based on the number of table rows."""
        font_size, leading, header_font_size, header_leading = 7, 9, 8, 11
        col_widths = [3.3*inch, 0.7*inch, 1.0*inch, 1.0*inch, 1.2*inch]
        header_padding, row_padding = 3, 1

        # Font setup (using original logic)
        font_name, font_name_bold = 'Jost', 'Jost-Bold'
        try:
            jost_regular_path = resource_path('Jost-Regular.ttf')
            jost_bold_path = resource_path('Jost-Bold.ttf')
            pdfmetrics.registerFont(TTFont(font_name, jost_regular_path))
            pdfmetrics.registerFont(TTFont(font_name_bold, jost_bold_path))
        except Exception as e:
            print(f"Could not load custom fonts, falling back to Helvetica. Error: {e}")
            font_name, font_name_bold = 'Helvetica', 'Helvetica-Bold'

        styles = getSampleStyleSheet()
        table_styles = {
            'header_project': ParagraphStyle('header_project_style', parent=styles['Normal'], fontName=font_name_bold, fontSize=14, alignment=0),
            'table_text': ParagraphStyle('table_text_style', parent=styles['Normal'], fontName=font_name, fontSize=font_size, leading=leading, alignment=0),
            'table_bold': ParagraphStyle('table_bold_style', parent=styles['Normal'], fontName=font_name_bold, fontSize=font_size, leading=leading, alignment=0),
            'table_bold_white': ParagraphStyle('table_bold_white_style', parent=styles['Normal'], fontName=font_name_bold, fontSize=font_size, leading=leading, textColor=colors.white, alignment=0),
            'table_header_left': ParagraphStyle('table_header_style_left', parent=styles['Normal'], fontName=font_name_bold, fontSize=header_font_size, leading=header_leading, alignment=0, textColor=colors.white),
            'table_header_right': ParagraphStyle('table_header_style_right', parent=styles['Normal'], fontName=font_name_bold, fontSize=header_font_size, leading=header_leading, alignment=2, textColor=colors.white)
        }
        
        return {
            'styles': table_styles,
            'col_widths': col_widths,
            'header_padding': header_padding,
            'row_padding': row_padding
        }

    def _create_pdf_header(self, style_settings):

        from reportlab.lib.utils import ImageReader

        styles = style_settings['styles']
        col_widths = style_settings['col_widths']  # [Name, Days, Start, Finish, Price]
        assert len(col_widths) == 5, "Expected 5 table columns to match main table."

        # Left cell: Company + Project (column 0)
        left_para = Paragraph(
            f"<font color='#991f2b'>{self.company_name.get()}<br/><br/>{self.project_name.get()}</font>",
            styles['header_project'],
        )

        # Middle cell: Client logo spanning columns 1–2 (centered)
        mid = Paragraph("", styles['header_project'])
        if self.client_logo_path.get() and os.path.exists(self.client_logo_path.get()):
            try:
                mid = Image(self.client_logo_path.get(),
                            width=2.0*inch, height=1.0*inch, kind='proportional')
                mid.hAlign = 'CENTER'
            except Exception:
                mid = Paragraph("", styles['header_project'])

        # Right cell: Company logo spanning columns 3–4 (right-aligned, larger area)
        right = Paragraph("", styles['header_project'])
        if self.logo_path.get() and os.path.exists(self.logo_path.get()):
            try:
                img = ImageReader(self.logo_path.get())
                iw, ih = img.getSize()

                # Make the logo as large as possible within the last TWO columns
                max_w = float(col_widths[3] + col_widths[4])   # Finish + Price total width
                max_h = 0.85 * inch                            # adjust if you want it taller
                scale_w = max_w / float(iw)
                scale_h = max_h / float(ih)
                scale = min(scale_w, scale_h)                      # fit inside both limits

                w = float(iw) * scale
                h = float(ih) * scale

                right = Image(self.logo_path.get(), width=w, height=h, kind='proportional')
                right.hAlign = 'RIGHT'
            except Exception:
                right = Paragraph("", styles['header_project'])

        # Build header on the exact same 5-column grid
        # Cells: [left_para, mid, "", right, ""]
        hdr = Table(
            [[left_para, mid, "", right, ""]],
            colWidths=col_widths,
            hAlign='LEFT',  # starts at doc.leftMargin like the main table
        )

        hdr.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

            ('ALIGN',  (0, 0), (0, 0), 'LEFT'),      # left text
            ('ALIGN',  (1, 0), (2, 0), 'CENTER'),    # client logo block
            ('ALIGN',  (3, 0), (4, 0), 'RIGHT'),     # company logo block

            # Spans
            ('SPAN',   (1, 0), (2, 0)),              # client logo spans cols 1–2
            ('SPAN',   (3, 0), (4, 0)),              # company logo spans cols 3–4

            # Paddings so edges align and logo hugs the right edge
            ('LEFTPADDING',  (0, 0), (0, 0), 0.15*inch),    # first column padding (match table)
            ('LEFTPADDING',  (1, 0), (2, 0), 0),
            ('RIGHTPADDING', (1, 0), (2, 0), 0),
            ('LEFTPADDING',  (3, 0), (4, 0), 0),
            ('RIGHTPADDING', (3, 0), (4, 0), -0.2*inch),  # tiny negative nudge to sit flush; use 0 if you prefer

            ('TOPPADDING',   (0, 0), (-1, -1), 15),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 0),
        ]))

        return hdr

    def _create_table_data(self, styles):
        """Prepare the data for the main project table."""
        all_table_data = []
        table_text_style = styles['table_text']
        table_bold_style = styles['table_bold']
        table_bold_white_style = styles['table_bold_white']
        table_header_style_left = styles['table_header_left']
        table_header_style_right = styles['table_header_right']

        # Header row
        header_row_formatted = [
            Paragraph('Project Milestones', table_header_style_left),
            Paragraph('Days', table_header_style_left),
            Paragraph('Start', table_header_style_left),
            Paragraph('Finish', table_header_style_left),
            Paragraph('Price', table_header_style_right)
        ]
        all_table_data.append(header_row_formatted)
        
        # Summary row
        total_price = sum(item.price for item in self.template_items if item.enabled.get() and item.indent_level == 0)
        valid_dates = [datetime.strptime(dt, "%m/%d/%y") for item in self.template_items if item.enabled.get() for dt in (item.start_date, item.end_date) if dt]
        earliest_start = min(valid_dates).strftime("%m/%d/%y") if valid_dates else ""
        latest_end = max(valid_dates).strftime("%m/%d/%y") if valid_dates else ""
        total_duration = self._get_business_days_between(earliest_start, latest_end)
        
        summary_row_formatted = [
            Paragraph(f"<b>{self.project_name.get()}</b>", table_bold_white_style),
            Paragraph(f"{total_duration}", table_bold_white_style),
            Paragraph(earliest_start, table_bold_white_style),
            Paragraph(latest_end, table_bold_white_style),
            Paragraph(f"${total_price:,}", ParagraphStyle('summary_price', parent=table_bold_white_style, alignment=2)),
        ]
        all_table_data.append(summary_row_formatted)

        # Recursive function to build all rows
        def build_table_rows_recursive(items):
            for item in items:
                if item.enabled.get():
                    is_main_milestone = item.is_milestone and item.indent_level == 0
                    current_style = table_bold_white_style if is_main_milestone else table_bold_style if item.is_milestone else table_text_style
                    price_style = ParagraphStyle('price_style', parent=current_style, alignment=2)
                    name_para_style = current_style
                    
                    if is_main_milestone:
                        name_text = f"<b>{'&nbsp;' * 4 * item.indent_level}{item.name}</b>"
                        name_para_style = ParagraphStyle('main_milestone_name', parent=table_bold_white_style)
                    elif item.is_milestone:
                        name_text = f"<b>{'&nbsp;' * 4 * item.indent_level}{item.name}</b>"
                        name_para_style = ParagraphStyle('sub_milestone_name', parent=table_bold_style)
                    else:
                        name_text = f"{'&nbsp;' * 4 * item.indent_level}{item.name}"

                    name_para = Paragraph(name_text, name_para_style)
                    row_data = [
                        name_para,
                        Paragraph(f"{item.duration}", current_style),
                        Paragraph(item.start_date, current_style),
                        Paragraph(item.end_date, current_style),
                        Paragraph(f"${item.price:,}" if item.price > 0 else ("$0" if item.is_milestone else ""), price_style),
                    ]
                    all_table_data.append(row_data)
                    
                    if item.children:
                        build_table_rows_recursive(item.children)
        
        build_table_rows_recursive(self.template_items)
        return all_table_data

    def _style_table(self, full_table, styles, style_settings):
        """Apply styles to the main project table."""
        table_style_commands = [
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,0), style_settings['header_padding']), 
            ('BOTTOMPADDING', (0,0), (-1,0), style_settings['header_padding']),
            ('TOPPADDING', (0,1), (-1,-1), style_settings['row_padding']),
            ('BOTTOMPADDING', (0,1), (-1,-1), style_settings['row_padding']),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#991f2b")),
            ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.black),
            ('BACKGROUND', (0, 1), (-1, 1), colors.black),
            ('LINEBELOW', (0, 1), (-1, 1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('LEFTPADDING',  (0, 0), (0, -1), 4),   # first column padding = 4pt (matches header)
            ('RIGHTPADDING', (-1, 0), (-1, -1), 4), # last column padding = 4pt (matches header)
        ]
        
        row_idx_offset = 2
        def find_and_style_milestones(items, current_row_idx):
            for item in items:
                if item.enabled.get():
                    if item.is_milestone:
                        bg_color = colors.HexColor("#991f2b") if item.indent_level == 0 else colors.HexColor("#D3D3D3")
                        table_style_commands.append(('BACKGROUND', (0, current_row_idx), (-1, current_row_idx), bg_color))
                    current_row_idx += 1
                    if item.children:
                        current_row_idx = find_and_style_milestones(item.children, current_row_idx)
            return current_row_idx

        find_and_style_milestones(self.template_items, row_idx_offset)
        full_table.setStyle(TableStyle(table_style_commands))
        return full_table

    def _add_gantt_page(self, elements, styles):
        """
        Collects data and adds a full-page, vector-based Gantt chart to the PDF story.
        """
        if not self.include_gantt.get():
            return

        # --- 1. Collect and prepare task data ---
        tasks_for_chart = []
        def collect_tasks_recursive(items):
            for item in items:
                if item.enabled.get() and not item.is_milestone and item.start_date and item.end_date and item.duration >= 0:
                    try:
                        start_dt = datetime.strptime(item.start_date, "%m/%d/%y")
                        end_dt = datetime.strptime(item.end_date, "%m/%d/%y")
                        tasks_for_chart.append({
                            "name": self.get_item_path(item),
                            "start": start_dt,
                            "end": end_dt
                        })
                    except (ValueError, TypeError):
                        continue # Skip tasks with invalid dates
                if item.children:
                    collect_tasks_recursive(item.children)

        collect_tasks_recursive(self.template_items)

        if not tasks_for_chart:
            return # Don't add a blank page if there's no data

        # --- 2. Determine the overall project date range for the timeline ---
        project_start_date = min(t['start'] for t in tasks_for_chart)
        project_end_date = max(t['end'] for t in tasks_for_chart)

        # --- 3. Add the Gantt chart page to the PDF elements ---
        elements.append(NextPageTemplate('LandscapePage'))
        elements.append(PageBreak())

        # Add a title for the schedule page
        title_style = ParagraphStyle('gantt_title', parent=styles['h1'], alignment=1, spaceAfter=0.1*inch)
        elements.append(Paragraph("Project Schedule", title_style))
        
        # Instantiate and add the custom Gantt Chart Flowable
        gantt_flowable = self.GanttChartFlowable(tasks_for_chart, project_start_date, project_end_date)
        elements.append(gantt_flowable)

    def create_pdf(self, filename):
        """
        MODIFIED: Create the multi-page PDF document using a modular approach.
        """
        doc = BaseDocTemplate(filename, topMargin=0.5*inch, bottomMargin=0.4*inch, leftMargin=0.3*inch, rightMargin=0.3*inch)
        
        portrait_frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='portrait_frame')
        l_width, l_height = landscape(letter)
        landscape_frame = Frame(doc.leftMargin, doc.bottomMargin,
                                l_width - doc.leftMargin - doc.rightMargin,
                                l_height - doc.bottomMargin - doc.topMargin,
                                id='landscape_frame')
        def _draw_footer(canv, _doc):
            canv.saveState()
            date_str = datetime.now().strftime("%B %d, %Y")
            version  = (self.version.get() or "V1").strip()
            text     = f"{date_str} - {version}"
            canv.setFont("Helvetica", 8)       # built-in for reliability
            y = 0.08 * inch                      # ultra-low; still print-safe
            x = doc.leftMargin+0.3*inch          # aligns to table's left edge
            canv.drawString(x, y, text)
            canv.restoreState()
        doc.addPageTemplates([
            PageTemplate(id='PortraitPage',  frames=[portrait_frame],  pagesize=letter,          onPage=_draw_footer),
            PageTemplate(id='LandscapePage', frames=[landscape_frame], pagesize=landscape(letter), onPage=_draw_footer),
    ])

        # Step 1: Count rows to determine dynamic table sizing
        def count_enabled_items(items):
            count = 0
            for item in items:
                if item.enabled.get():
                    count += 1
                    if item.children:
                        count += count_enabled_items(item.children)
            return count
        
        num_rows = count_enabled_items(self.template_items) + 2

        # Step 2: Setup styles and table properties dynamically
        style_settings = self._setup_reportlab_styles(num_rows)
        
        elements = []
        
        # Step 3: Add the header to the first page
        elements.append(self._create_pdf_header(style_settings))

        elements.append(Spacer(1, 0.2*inch))

        # Step 4: Create and style the main project table
        table_data = self._create_table_data(style_settings['styles'])
        full_table = Table(table_data, colWidths=style_settings['col_widths'], repeatRows=1)
        full_table = self._style_table(full_table, style_settings['styles'], style_settings)
        elements.append(full_table)
        
        # Step 5: Add the Gantt chart page if requested
        self._add_gantt_page(elements, getSampleStyleSheet())
        
        doc.build(elements)

    # --- MODIFICATION: Replaced save_template with save_template_excel ---
    def save_template_excel(self):
        """Save current template to a cleanly formatted Excel file."""
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save Template As"
        )
        if not filename:
            return

        try:
            wb = Workbook()
            # Sheet 1: Project Info
            info_ws = wb.active
            info_ws.title = "Project Info"
            
            logo_path_to_save = self.logo_path.get()
            if logo_path_to_save == self.default_logo_path:
                logo_path_to_save = "DEFAULT_LOGO"

            project_data = {
                "Project Name": self.project_name.get(),
                "Company Name": self.company_name.get(),
                "Project Start Date": self.project_start_date.get(),
                "Logo Path": logo_path_to_save,
                "Client Logo Path": self.client_logo_path.get(),
                "Last Task ID": self.task_counter
            }

            info_ws.append(["Attribute", "Value"])
            for key, value in project_data.items():
                info_ws.append([key, value])
            
            # Style the info sheet
            info_ws['A1'].font = Font(bold=True)
            info_ws['B1'].font = Font(bold=True)
            info_ws.column_dimensions['A'].width = 20
            info_ws.column_dimensions['B'].width = 50

            # Sheet 2: Tasks
            tasks_ws = wb.create_sheet(title="Tasks")
            headers = [
                "ID", "Name", "Duration", "Price", "Is Milestone", "Indent Level",
                "Enabled", "Predecessor ID", "Lag", "Is Start Pinned", "Parent ID"
            ]
            tasks_ws.append(headers)
            for cell in tasks_ws[1]:
                cell.font = Font(bold=True)

            # Flatten the hierarchical data
            flat_tasks = []
            def _flatten_items(items, parent_id=None):
                for item in items:
                    flat_tasks.append({
                        "ID": item.id, "Name": item.name, "Duration": item.duration,
                        "Price": item.price, "Is Milestone": item.is_milestone,
                        "Indent Level": item.indent_level, "Enabled": item.enabled.get(),
                        "Predecessor ID": item.predecessor_id, "Lag": item.lag,
                        "Is Start Pinned": item.is_start_pinned, "Parent ID": parent_id
                    })
                    if item.children:
                        _flatten_items(item.children, parent_id=item.id)
            
            _flatten_items(self.template_items)

            # Write tasks to sheet
            for task in flat_tasks:
                row = [task.get(h) for h in headers]
                tasks_ws.append(row)
            
            # Auto-size columns for tasks sheet
            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                tasks_ws.column_dimensions[column_letter].width = max(15, len(header) + 2)


            wb.save(filename)
            messagebox.showinfo("Success", "Excel template saved successfully!")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel template: {str(e)}")

    # --- MODIFICATION: Replaced load_template with load_template_excel ---
    def load_template_excel(self):
        """Load a template from an Excel file."""
        filename = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx")],
            title="Load Excel Template"
        )
        if not filename:
            return

        try:
            wb = load_workbook(filename)
            
            # Load Project Info
            info_ws = wb["Project Info"]
            project_data = {row[0]: row[1] for row in info_ws.iter_rows(min_row=2, values_only=True)}

            self.project_name.set(project_data.get("Project Name", ""))
            self.company_name.set(project_data.get("Company Name", ""))
            self.project_start_date.set(project_data.get("Project Start Date", ""))
            self.task_counter = int(project_data.get("Last Task ID", 0))

            saved_logo_path = project_data.get("Logo Path", "")
            if saved_logo_path == "DEFAULT_LOGO" or not os.path.exists(saved_logo_path):
                self.logo_path.set(self.default_logo_path)
            else:
                self.logo_path.set(saved_logo_path)
            self.client_logo_path.set(project_data.get("Client Logo Path", ""))

            # Load Tasks and reconstruct hierarchy
            tasks_ws = wb["Tasks"]
            headers = [cell.value for cell in tasks_ws[1]]
            
            items_by_id = {}
            all_items_data = []

            for row in tasks_ws.iter_rows(min_row=2, values_only=True):
                item_data = dict(zip(headers, row))
                all_items_data.append(item_data)
                
                item_id = int(item_data["ID"])
                item = ProposalItem(
                    name=item_data["Name"],
                    duration=math.ceil(float(item_data.get("Duration", 0))),
                    price=int(item_data.get("Price", 0)),
                    is_milestone=bool(item_data.get("Is Milestone")),
                    indent_level=int(item_data.get("Indent Level", 0)),
                    item_id=item_id
                )
                item.enabled.set(bool(item_data.get("Enabled", True)))
                pred_id = item_data.get("Predecessor ID")
                item.predecessor_id = int(pred_id) if pred_id is not None else None
                item.lag = int(item_data.get("Lag", 0))
                item.is_start_pinned = bool(item_data.get("Is Start Pinned", False))
                
                items_by_id[item_id] = item

            # Rebuild the tree structure
            root_items = []
            for item_data in all_items_data:
                item_id = int(item_data["ID"])
                parent_id_val = item_data.get("Parent ID")
                parent_id = int(parent_id_val) if parent_id_val is not None else None
                
                current_item = items_by_id[item_id]
                
                if parent_id in items_by_id:
                    parent_item = items_by_id[parent_id]
                    parent_item.children.append(current_item)
                    current_item.parent = parent_item
                else:
                    root_items.append(current_item)

            self.template_items = root_items
            self.populate_tree()
            self.expand_all_items()
            messagebox.showinfo("Success", "Excel template loaded successfully!")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel template: {str(e)}")

# Your existing UI/engine
# from proposal import ProposalGenerator, ProposalItem  # merged above
# ===================
# Parsing & Build Rules
# ===================

EXACT_SUBTOTAL_LABELS = {
    "civil engineering", "electrical engineering", "structural engineering", "substation engineering"
}

PHASES = ["30%", "60%", "90%", "IFC"]


def _pairs(ncols: int):
    """(task, price) column pairs on Proposal Page. Includes D/E explicitly."""
    return [(0, 1), (3, 4), (6, 7), (9, 10), (10, 11)]


def _categorize(text: str) -> str:
    t = (text or "").lower().strip()
    if t.startswith("civil"): return "Civil"
    if t.startswith("electrical"): return "Electrical"
    if t.startswith("structural"): return "Structural"
    if t.startswith("substation"): return "Substation"
    return ""


def _infer_phase(text: str):
    t = (text or "").lower()
    if "30% design" in t or "30%" in t: return "30%"
    if "60% design" in t or "60%" in t: return "60%"
    if "90% design" in t or "90%" in t: return "90%"
    if "ifc design" in t or "record drawings" in t or "ifc" in t: return "IFC"
    return None


def load_proposal_page_rows(path: str):
    pp = pd.read_excel(path, sheet_name="Proposal Page", header=None, engine="openpyxl")
    props = []
    ncols = pp.shape[1]

    def normalize(s):
        return (s or "").strip()

    for (txt_col, price_col) in _pairs(ncols):
        current_phase = None
        current_category = None  # NEW

        for i in range(pp.shape[0]):
            txt = pp.iat[i, txt_col] if txt_col < ncols else None
            val = pp.iat[i, price_col] if price_col < ncols else None

            # Phase detection
            if isinstance(txt, str):
                maybe = _infer_phase(txt)
                if maybe:
                    current_phase = maybe

                # Category header detection (NEW)
                low = txt.lower().strip()
                if low in ("civil engineering", "electrical engineering", "structural engineering", "substation engineering"):
                    current_category = txt.split()[0].capitalize()  # "Civil", "Electrical", etc.
                    continue  # category header row itself isn't a task

            # Candidate task
            if isinstance(txt, str) and normalize(txt) and pd.notna(val):
                lower = txt.lower().strip()
                if any(k in lower for k in [
                    "total", "milestone", "summary of services", "engineering proposal", "insurance adder"
                ]):
                    continue
                if lower in EXACT_SUBTOTAL_LABELS:
                    continue

                # price
                try:
                    price = float(val)
                except Exception:
                    continue

                # Category assignment:
                cat = _categorize(txt) or current_category or ""
                if not cat:
                    # If still unknown, skip as before
                    continue

                props.append({
                    "category": cat,
                    "task": normalize(txt),
                    "proposal_price": price,
                    "phase": current_phase or _infer_phase(txt) or "30%",
                })
    return props



def _load_detail_map(path: str, sheet: str):
    """Return {Description: {hours, price}}; dedupe by keeping first priced/max price."""
    try:
        df = pd.read_excel(path, sheet_name=sheet, header=None, engine="openpyxl")
    except Exception:
        return {}
    df2 = df.iloc[12:].reset_index(drop=True)  # after header row (index 11)
    DESC, HRS, COST = 2, 11, 12
    out = {}
    for _, row in df2.iterrows():
        desc = row.iloc[DESC] if len(row) > DESC else None
        hours = row.iloc[HRS] if len(row) > HRS else None
        price = row.iloc[COST] if len(row) > COST else None
        if isinstance(desc, str) and desc.strip():
            name = desc.strip()
            h = float(hours) if pd.notna(hours) else 0.0
            p = 0.0
            if pd.notna(price) and str(price).strip().lower() != "not included":
                try:
                    p = float(price)
                except Exception:
                    p = 0.0
            if name not in out:
                out[name] = {"hours": h, "price": p}
            else:
                prev_p = float(out[name].get("price") or 0.0)
                if (prev_p <= 0 and p > 0) or (p > prev_p):
                    out[name] = {"hours": h, "price": p}
    return out


def _load_structural_from_electrical(path: str):
    """Structural hours/prices live under 'Structural Engineering' section in Electrical sheet."""
    try:
        df = pd.read_excel(path, sheet_name="Electrical", header=None, engine="openpyxl")
    except Exception:
        return {}
    DESC, HRS, COST = 2, 11, 12

    start_idx = None
    for i in range(len(df)):
        if any(isinstance(v, str) and "structural engineering" in v.lower() for v in df.iloc[i].tolist()):
            start_idx = i + 1
            break
    if start_idx is None:
        return {}

    out = {}
    blank_streak = 0
    for i in range(start_idx, len(df)):
        desc = df.iat[i, DESC] if DESC < df.shape[1] else None
        hours = df.iat[i, HRS] if HRS < df.shape[1] else None
        price = df.iat[i, COST] if COST < df.shape[1] else None

        stop_here = False
        if isinstance(desc, str):
            low = desc.lower().strip()
            if (low.endswith("engineering") and low != "structural engineering") or " design" in low or "—" in desc or " - " in desc:
                stop_here = True

        if not isinstance(desc, str) or not desc.strip():
            blank_streak += 1
            if blank_streak >= 2:
                break
            if stop_here:
                break
            continue
        else:
            blank_streak = 0
            if stop_here:
                break

        name = desc.strip()
        h = float(hours) if pd.notna(hours) else 0.0
        pval = 0.0
        if pd.notna(price) and str(price).strip().lower() != "not included":
            try:
                pval = float(price)
            except Exception:
                pval = 0.0
        if name not in out or pval > float(out[name].get("price") or 0):
            out[name] = {"hours": h, "price": pval}
    return out


def enrich_with_details(path: str, rows):
    xl = pd.ExcelFile(path, engine="openpyxl")
    civil_map = _load_detail_map(path, "Civil") if "Civil" in xl.sheet_names else {}
    elec_map = _load_detail_map(path, "Electrical") if "Electrical" in xl.sheet_names else {}
    structural_from_elec = _load_structural_from_electrical(path)

    for r in rows:
        if r["category"] == "Electrical":
            d = elec_map.get(r["task"], {})
        elif r["category"] == "Civil":
            d = civil_map.get(r["task"], {})
        elif r["category"] == "Structural":
            d = structural_from_elec.get(r["task"], {})
        else:
            d = {}
        r["hours"] = float(d.get("hours")) if d.get("hours") is not None else None
        r["detail_price"] = float(d.get("price")) if d.get("price") is not None else None
    return rows


def extract_project_info(path: str):
    df = pd.read_excel(path, sheet_name="Proposal Page", header=None, engine="openpyxl")
    info = {"date": None, "client": None, "project": None, "location": None}
    try:
        info["date"] = df.iat[0, 1]
        info["client"] = df.iat[1, 1] if isinstance(df.iat[1, 0], str) and "client" in str(df.iat[1, 0]).lower() else None
        info["project"] = df.iat[2, 1] if isinstance(df.iat[2, 0], str) and "project" in str(df.iat[2, 0]).lower() else None
        info["location"] = df.iat[0, 4] if isinstance(df.iat[0, 3], str) and "location" in str(df.iat[0, 3]).lower() else None
    except Exception:
        pass
    return info


def build_model_rows(path: str):
    rows = load_proposal_page_rows(path)
    rows = enrich_with_details(path, rows)

    # Normalize phases
    for r in rows:
        r["phase"] = r.get("phase") or _infer_phase(r["task"]) or "30%"

    def is_electrical_study(name: str) -> bool:
        n = (name or "").lower()
        return "study" in n and (n.startswith("electrical") or "electrical study" in n)

    # Inclusion filter
    # Inclusion filter (RELAXED):
    filtered = []
    for r in rows:
        cat = r["category"]
        pp = r.get("proposal_price") or 0
        dp = r.get("detail_price") or 0

        keep = False
        # If either proposal or detail price is present, keep it.
        if pp > 0 or dp > 0:
            keep = True

        # (Optional) for Electrical "study" heuristic can still be kept, but is redundant now
        if keep:
            filtered.append(r)
    rows = filtered


    # Bucket by category/phase
    buckets = {
        "Civil": {p: [] for p in PHASES},
        "Electrical": {p: [] for p in PHASES},
        "Structural": {p: [] for p in PHASES},
    }
    for r in rows:
        cat, ph = r["category"], r["phase"]
        if cat in buckets and ph in buckets[cat]:
            buckets[cat][ph].append(r)

    info = extract_project_info(path)
    return buckets, info


def flatten_to_template_rows(buckets, hours_per_day: float, price_source: str, review_pairs: set):
    """
    Flatten into the table format ProposalGenerator expects (via ProposalItem build below).
    - Project Initiation (w/ Civil/Electrical Due Diligence 1d)
    - Civil, Electrical, Structural (30/60/90/IFC) with rules:
        * 30% & 60% phases get Client Review (always on)
        * First task of 30% depends on matching Due Diligence (Civil/Electrical)
        * Durations from hours (Civil/Electrical), Structural 0d
    - Prices:
        * "proposal" → Proposal Page price
        * "detail"   → Detail price (fallback to Proposal)
    - Project Closeout top-level milestone only (empty, 0d)
    """
    rows_out = []
    next_id = 1

    # NEW: cross-category predecessor trackers
    e60_first_task_id = None               # first task ID of Electrical 60%
    structural_first_task_applied = False  # ensure we only set Structural's first task once

    def add_item(item_id, name, duration, price, is_milestone, indent, enabled, pred_id, lag, pinned, parent_id):
        rows_out.append({
            "ID": item_id,
            "Name": name,
            "Duration": int(duration or 0),
            "Price": int(price or 0),
            "Is Milestone": bool(is_milestone),
            "Indent Level": int(indent),
            "Enabled": bool(enabled),
            "Predecessor ID": pred_id,
            "Lag": int(lag or 0),
            "Is Start Pinned": bool(pinned),
            "Parent ID": parent_id,
        })

    # ---- Project Initiation ----
    pi_id = next_id; next_id += 1
    add_item(pi_id, "Project Initiation", 0, 0, True, 0, True, None, 0, False, None)

    last_pi_child = None
    civil_dd_id = None
    electrical_dd_id = None
    for name, dur in [
        ("Deposit & Contract Signed", 0),
        ("Notice to Proceed", 0),
        ("Civil Start - Civil Due Diligence", 1),
        ("Electrical Start - Electrical Due Diligence", 1),
    ]:
        tid = next_id; next_id += 1
        add_item(tid, name, dur, 0, False, 1, True, last_pi_child, 0, False, pi_id)
        last_pi_child = tid
        if name.startswith("Civil Start"): civil_dd_id = tid
        if name.startswith("Electrical Start"): electrical_dd_id = tid

    def build_category(cat_key, cat_label):
        nonlocal next_id, e60_first_task_id, structural_first_task_applied
        # Create the top-level milestone for the category, disabled by default.
        cat_id = next_id; next_id += 1
        add_item(cat_id, cat_label, 0, 0, True, 0, False, None, 0, False, None)

        added_any = False  # only enable the top-level if we actually add tasks/milestones

        def _price_of(t):
            # Price selection policy by source, with Structural special-casing preserved.
            if cat_key == "Structural":
                if price_source == "detail":
                    return (t.get("detail_price") or t.get("proposal_price") or 0)
                return (t.get("proposal_price") or t.get("detail_price") or 0)
            if price_source == "detail":
                return (t.get("detail_price") or t.get("proposal_price") or 0)
            return (t.get("proposal_price") or t.get("detail_price") or 0)

        def _reorder_for_30(tasks, phase):
            # Nudge the "plan set" to the top at 30% like your previous heuristic.
            if phase != "30%":
                return list(tasks)
            def score(t): 
                return 0 if "plan set" in (t.get("task", "").lower()) else 1
            return sorted(tasks, key=score)

        prev_phase_last_id = None  # last task in the previous phase within this category

        for phase in PHASES:
            raw = buckets[cat_key][phase]
            if not raw:
                continue

            added_any = True  # we will enable the category since this phase has content
            tasks = _reorder_for_30(raw, phase)

            # Phase milestone (indented under the category)
            ms_id = next_id; next_id += 1
            add_item(ms_id, f"{cat_label} — {phase} Design", 0, 0, True, 1, True, None, 0, False, cat_id)

            first_task_id = None
            last_task_id = None

            for t in tasks:
                # Duration: Civil/Electrical from hours; Structural = 0d
                if cat_key in ("Civil", "Electrical","Structural") and t.get("hours") is not None:
                    dur_days = math.ceil((t.get("hours") or 0) / float(hours_per_day))
                else:
                    dur_days = 0

                tid = next_id; next_id += 1
                add_item(
                    tid,
                    t["task"],
                    dur_days,
                    _price_of(t),
                    False,          # is_milestone
                    2,              # indent under the phase milestone
                    True,           # enabled
                    last_task_id,   # predecessor (sequential within the phase)
                    0,              # lag
                    False,          # pinned
                    ms_id           # parent is phase milestone
                )
                if first_task_id is None:
                    first_task_id = tid
                last_task_id = tid

            # Client Review (10d, $0) for the selected pairs
            if (cat_key, phase) in review_pairs:
                cr_id = next_id; next_id += 1
                add_item(cr_id, "Client Review", 10, 0, False, 2, True, last_task_id, 0, False, ms_id)
                last_task_id = cr_id

            # Capture the first task of Electrical 60% to link Structural later
            if cat_key == "Electrical" and phase == "60%" and first_task_id is not None and e60_first_task_id is None:
                e60_first_task_id = first_task_id
            elif cat_key == "Electrical" and phase == "90%" and first_task_id is not None and e60_first_task_id is None:
                e60_first_task_id = first_task_id

            # Set cross-phase/intro predecessors for the *first* task in this phase
            if first_task_id is not None:
                # Find the row to edit by ID
                idx = next(i for i in range(len(rows_out)) if rows_out[i]["ID"] == first_task_id)

                if cat_key == "Structural" and (not structural_first_task_applied) and e60_first_task_id:
                    # Structural’s very first task depends on the first task of Electrical 60%
                    rows_out[idx]["Predecessor ID"] = e60_first_task_id
                    structural_first_task_applied = True
                elif prev_phase_last_id is not None:
                    # Otherwise, chain the first task of this phase to the last task of the previous phase (same category)
                    rows_out[idx]["Predecessor ID"] = prev_phase_last_id
                else:
                    # For the first phase of each category, wire to Due Diligence (if 30%) or to last Project Initiation child
                    if phase == "30%":
                        if cat_key == "Civil" and civil_dd_id:
                            rows_out[idx]["Predecessor ID"] = civil_dd_id
                        elif cat_key == "Electrical" and electrical_dd_id:
                            rows_out[idx]["Predecessor ID"] = electrical_dd_id
                        else:
                            rows_out[idx]["Predecessor ID"] = last_pi_child
                    else:
                        rows_out[idx]["Predecessor ID"] = last_pi_child

            # Remember the boundary of this phase to chain the next phase’s first task
            prev_phase_last_id = last_task_id

        # Finally, enable the top-level category milestone iff we actually added content.
        if added_any:
            for i in range(len(rows_out) - 1, -1, -1):
                if rows_out[i]["ID"] == cat_id:
                    rows_out[i]["Enabled"] = True
                    break

        return cat_id


    # Default (non-optional) review pairs per spec
    review_pairs = {("Civil", "30%"), ("Civil", "60%"), ("Electrical", "30%"), ("Electrical", "60%")}

    # Order: Civil → Electrical → Structural
    build_category("Civil", "Civil Engineering")
    build_category("Electrical", "Electrical Engineering")
    build_category("Structural", "Structural Engineering")

    # ---- Project Closeout (top-level only, empty, 0d) ----
    closeout_id = next_id; next_id += 1
    add_item(closeout_id, "Project Closeout", 0, 0, True, 0, True, None, 0, False, None)

    return rows_out



def push_into_generator(gen: ProposalGenerator, project_info, rows_out):
    """Replace any existing task tree with the new one and refresh the UI."""
    # Try to clear any existing Treeview if present
    try:
        tree = getattr(gen, "tree", None) or getattr(gen, "treeview", None)
        if tree is not None and hasattr(tree, "get_children"):
            for iid in tree.get_children(""):
                tree.delete(iid)
    except Exception:
        pass

    # Reset internal containers
    gen.template_items = []
    gen.item_id_map = {}
    gen.task_counter = 0

    # Rebuild from rows
    id_to_item = {}

    def mk_item(row):
        return ProposalItem(
            name=row["Name"],
            duration=int(row["Duration"] or 0),
            price=int(row["Price"] or 0),
            is_milestone=bool(row["Is Milestone"]),
            indent_level=int(row["Indent Level"] or 0),
            item_id=int(row["ID"]),
        )

    for r in rows_out:
        id_to_item[r["ID"]] = mk_item(r)

    for r in rows_out:
        item = id_to_item[r["ID"]]
        pid = r["Parent ID"]
        if pid:
            parent = id_to_item.get(pid)
            if parent:
                item.parent = parent
                parent.children.append(item)
        pred = r["Predecessor ID"]
        if pred:
            item.predecessor_id = int(pred)
            item.predecessor_type = 'FS'
            item.lag = int(r.get("Lag") or 0)
        item.enabled.set(bool(r["Enabled"]))

    gen.template_items = [it for it in id_to_item.values() if it.parent is None]
    gen.item_id_map = {it.id: it for it in id_to_item.values()}
    gen.task_counter = max(gen.item_id_map.keys()) if gen.item_id_map else 0

    # Project info
    if project_info.get("project"):
        gen.project_name.set(str(project_info["project"]))
    if project_info.get("client"):
        gen.company_name.set(str(project_info["client"]))
    date_cell = project_info.get("date")
    try:
        if isinstance(date_cell, (pd.Timestamp, datetime)):
            dt = pd.to_datetime(date_cell).to_pydatetime()
            gen.project_start_date.set(dt.strftime("%m/%d/%y"))
        elif isinstance(date_cell, str) and date_cell.strip():
            dt = pd.to_datetime(date_cell)
            gen.project_start_date.set(pd.to_datetime(dt).strftime("%m/%d/%y"))
    except Exception:
        pass

    # If ProposalGenerator exposes a UI refresh method, call it
    for meth in ("rebuild_tree", "refresh_tree", "render_tree", "draw_tree"):
        if hasattr(gen, meth):
            try:
                getattr(gen, meth)()
                break
            except Exception:
                pass

